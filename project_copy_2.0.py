import os
import shutil
from tkinter.filedialog import askdirectory
import tkinter
import openpyxl
from datetime import datetime

current_dir = os.getcwd()
excel_file = 'copy-data.xlsx'
excel_file_n_path = os.path.join(current_dir, excel_file)


wb = openpyxl.load_workbook(excel_file_n_path)
sheet = wb['Blad1']

root = tkinter.Tk()
root.withdraw()


def openWriteFile(new_file, from_source):
    now = datetime.now()
    time = now.strftime("%Y-%m-%d, %H:%M:%S")
    file = open("logfile.txt", "a+")
    file.write(time + ": "+"Created: " + new_file +
               " From: " + from_source + "\n")
    file.close()


def createFiles():
    # source_folder_n_path = askdirectory(title='--------Select Source Folder')
    folder_n_path = askdirectory(title='--------Select Destination Folder')

    try:
        makeFolder = os.makedirs(folder_n_path)
        print("Destination folder missing, creating...")
    except:
        print("Destination folder located, duplicating files...")

    counter = 0
    failCounter = 0

    # Logfile.txt
    now = datetime.now()
    time = now.strftime("%Y-%m-%d, %H:%M:%S")
    file = open("logfile.txt", "a+")
    file.write("-----------------------------------------\n")
    file.write("Job started at " + time + "\n")
    file.write("-----------------------------------------\n")
    file.close()

    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        for char in row:
            if type(char) is str:
                fileExtention = char[-4:]

        # Copying the primary picture and renaming it according the the Excel file.
        file_name_front = str(row[3])+"_01"+str(fileExtention)
        source_file_n_path_front = os.path.join(row[0], str(row[1]))
        file_n_path_destination_front = os.path.join(
            folder_n_path, file_name_front)
        if row[1] != None:
            shutil.copy(source_file_n_path_front,
                        file_n_path_destination_front)
            print(file_n_path_destination_front)
            counter += 1
            openWriteFile(str(file_n_path_destination_front),
                          str(source_file_n_path_front))
        else:
            failCounter += 1
            continue

        # Copying the secondary picture and renaming it according the the Excel file.
        file_name_back = str(row[3])+"_02"+str(fileExtention)
        source_file_n_path_back = os.path.join(row[0], str(row[2]))
        file_n_path_destination_back = os.path.join(
            folder_n_path, file_name_back)
        if row[2] != None:
            shutil.copy(source_file_n_path_back, file_n_path_destination_back)
            print(file_n_path_destination_back)
            counter += 1
            openWriteFile(str(file_n_path_destination_back),
                          str(source_file_n_path_back))
        else:
            failCounter += 1
            continue

    file = open("logfile.txt", "a+")
    file.write("Job complete " + time + "\n" +
               str(counter) + " files copied.\r\n")
    file.close()

    print("Done!")
    print(counter, "files duplicated and renamed!")
    print(failCounter, "empty cell in Excel file, therefore no action")


def preCheck():
    print("Checking source files vs Excel file...")

    missing_files = []
    path_list = []
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        for pic in range(1, 2):
            path_list.append(row[pic])

        files_in_dir = os.listdir(row[0])
        # print(files_in_dir)
        for i in path_list:
            if i not in files_in_dir and not i == None:
                missing_files.append(i)

    if len(missing_files) <= 0:
        print("All files accounted for! Proceeding...")
        createFiles()
    else:
        missing_files_set = set(missing_files)
        missing_files_unique = list(missing_files_set)
        print("Missing", len(missing_files_unique),
              "source files:", missing_files_unique)
        input()


print("Copy and Rename 2.0")
preCheck()
input()

# Man bör kunna göra allt detta med en loop, till och med med numpy eller panda.
