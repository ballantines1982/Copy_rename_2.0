import os
import shutil
from tkinter.filedialog import askdirectory
import tkinter as tk
from tkinter import ttk
import openpyxl
from datetime import datetime


class mainFrame(ttk.Frame):
    def __init__(self, container):
        super().__init__(container)

        self.preCheckPassed = False

        options = {'pady': 10, 'padx': 10}

        self.labelTitle = ttk.Label(self, text="Copy Rename 4.0 ADMIN")
        self.labelTitle.grid(column=0, row=0, sticky=tk.NW, **options)

        self.labelFrameGuide = ttk.LabelFrame(self, text="Instruktioner")
        self.labelFrameGuide.grid(
            column=0, row=1, columnspan=2, sticky=tk.EW, **options)

        self.labelInstruktion = ttk.Label(
            self.labelFrameGuide, text="1. Kontrollera Excelfilen\n2. Kontrollera datan\n3. Kör programmet")
        self.labelInstruktion.grid(
            column=0, row=0, columnspan=3, sticky=tk.NSEW, **options)

        self.btnExcel = ttk.Button(
            self.labelFrameGuide, text="Öppna Excelfilen")
        self.btnExcel['command'] = self.openExcelFile
        self.btnExcel.grid(column=0, row=1, sticky=tk.EW, padx=10, pady=10)
        self.btnCheck = ttk.Button(
            self.labelFrameGuide, text="Kontrollera data")
        self.btnCheck['command'] = self.preCheck
        self.btnCheck.grid(column=1, row=1, sticky=tk.EW, padx=40, pady=10)

        self.btnRun = ttk.Button(self.labelFrameGuide, text="Kör")
        if (self.preCheckPassed == False):
            self.btnRun['state'] = tk.DISABLED
        else:
            self.btnRun['state'] = tk.NORMAL
        self.btnRun['command'] = self.createFiles
        self.btnRun.grid(column=2, row=1, sticky=tk.EW, padx=10, pady=10)

        self.labelFrameStatus = ttk.LabelFrame(self, text="Status")
        self.labelFrameStatus.grid(column=0, row=2, columnspan=2, **options)
        self.statusText = tk.Text(self.labelFrameStatus, height=10, width=45)
        self.statusText.config(font=('verdana', 9))
        self.statusText.grid(column=0, row=2, sticky=tk.NS,
                             columnspan=2, **options)
        self.grid()

        self.current_dir = os.getcwd()
        self.excel_file = 'copy-data.xlsx'
        self.excel_file_n_path = os.path.join(
            self.current_dir, self.excel_file)

    def openWriteFile(self, new_file, from_source):
        now = datetime.now()
        time = now.strftime("%Y-%m-%d, %H:%M:%S")
        file = open("logfile.txt", "a+")
        file.write(time + ": "+"Created: " + new_file +
                   " From: " + from_source + "\n")
        file.close()

    def preCheck(self):
        self.statusText.delete("1.0", tk.END)
        try:
            self.wb = openpyxl.load_workbook(self.excel_file_n_path)
            self.sheet = self.wb['Blad1']
        except PermissionError:
            self.statusText.insert(
                tk.END, "ERROR: Vänligen stäng " + self.excel_file)
        except Exception as e:
            self.statusText.insert(tk.END, e)
        else:
            self.statusText.insert(
                tk.END, f'{self.excel_file} hittad! Fortsätter...\n')
        self.statusText.insert(
            tk.END, "Kontrollerar sökvägar i Excelfilen...\n")

        missing_files = []
        # picture_names = []
        path_list = []

        try:
            for row in self.sheet.iter_rows(min_row=2, max_col=5, values_only=True):

                for pic in range(1, 5):

                    if row[pic] != None:
                        picture_names = row[pic]
                        path_to_source_folder = row[0]
                        full_path = os.path.join(
                            path_to_source_folder, picture_names)
                        # path_list.append(full_path)
                        if os.path.isfile(full_path) != True:
                            missing_files.append(picture_names)

        except PermissionError:
            self.statusText.insert(
                tk.END, "ERROR: Vänligen stäng " + self.excel_file)
        except Exception as e:
            self.statusText.insert(tk.END, e)
        else:
            if len(missing_files) <= 0:
                self.statusText.insert(
                    tk.END, "Alla bilder hittade! \nKontroll OK!\n")
                self.btnRun['state'] = tk.NORMAL
            else:
                missing_files_set = set(missing_files)
                missing_files_unique = list(missing_files_set)
                self.statusText.insert(
                    tk.END, f'Saknar totalt {len(missing_files_unique)} bild/bilder:\n')
                for pic in missing_files_unique:
                    self.statusText.insert(
                        tk.END, f'{pic}\n')
                        

    def createFiles(self):
        folder_n_path = askdirectory(title='--------Välj Folder för bilderna')
        if folder_n_path:

            try:
                makeFolder = os.makedirs(folder_n_path)
                self.statusText.insert(
                    tk.END, "\nMappen saknas, skapar mapp...")
            except:
                self.statusText.insert(
                    tk.END, "Mappen hittad, fortsätter...\n")

            counter = 0
            failCounter = 0

            # Logfile.txt
            now = datetime.now()
            time = now.strftime("%Y-%m-%d, %H:%M:%S")
            file = open("logfile.txt", "a+")
            file.write("-----------------------------------------\n")
            file.write("Job started at " + time + "\n")
            file.write("-----------------------------------------\n")
            file.close()

            try:
                for row in self.sheet.iter_rows(min_row=2, max_col=6, values_only=True):
                    for char in row:
                        if isinstance(char, str):
                            if ".jpg" in char:
                                fileExtention = char[-4:]

                            if ".jpeg" in char:
                                fileExtention = char[-5:]

                    # Copying the primary picture and renaming it according the the Excel file.
                    file_name_front = str(row[5])+"_h1"+str(fileExtention)
                    source_file_n_path_front = os.path.join(
                        row[0], str(row[1]))
                    file_n_path_destination_front = os.path.join(
                        folder_n_path, file_name_front)
                    if row[1] != None:
                        shutil.copy(source_file_n_path_front,
                                    file_n_path_destination_front)
                        # Adding a dot (.) in the UI
                        self.statusText.insert(tk.END, '.')
                        counter += 1
                        # Logging
                        self.openWriteFile(str(file_n_path_destination_front),
                                           str(source_file_n_path_front))
                    else:
                        failCounter += 1
                        continue

                    # Copying the secondary picture and renaming it according the the Excel file.
                    file_name_back = str(row[5])+"_h2"+str(fileExtention)
                    source_file_n_path_back = os.path.join(row[0], str(row[2]))
                    file_n_path_destination_back = os.path.join(
                        folder_n_path, file_name_back)
                    if row[2] != None:
                        shutil.copy(source_file_n_path_back,
                                    file_n_path_destination_back)
                        # Adding a dot (.) in the UI
                        self.statusText.insert(tk.END, '.')
                        counter += 1
                        # Logging
                        self.openWriteFile(str(file_n_path_destination_back),
                                           str(source_file_n_path_back))
                    else:
                        failCounter += 1
                        continue

                    # Copying the primary picture and renaming it according the the Excel file.
                    file_name_third = str(row[5])+"_h3"+str(fileExtention)
                    source_file_n_path_third = os.path.join(
                        row[0], str(row[3]))
                    file_n_path_destination_third = os.path.join(
                        folder_n_path, file_name_third)
                    if row[3] != None:
                        shutil.copy(source_file_n_path_third,
                                    file_n_path_destination_third)
                        # Adding a dot (.) in the UI
                        self.statusText.insert(tk.END, '.')
                        counter += 1
                        # Logging
                        self.openWriteFile(str(file_n_path_destination_third),
                                           str(source_file_n_path_third))
                    else:
                        failCounter += 1
                        continue
                    
                    # Copying the primary picture and renaming it according the the Excel file.
                    file_name_forth = str(row[5])+"_h4"+str(fileExtention)
                    source_file_n_path_forth = os.path.join(
                        row[0], str(row[4]))
                    file_n_path_destination_forth = os.path.join(
                        folder_n_path, file_name_forth)
                    if row[4] != None:
                        shutil.copy(source_file_n_path_forth,
                                    file_n_path_destination_forth)
                        # Adding a dot (.) in the UI
                        self.statusText.insert(tk.END, '.')
                        counter += 1
                        # Logging
                        self.openWriteFile(str(file_n_path_destination_forth),
                                           str(source_file_n_path_forth))
                    else:
                        failCounter += 1
                        continue
                    
                    
                file = open("logfile.txt", "a+")
                file.write("Job complete " + time + "\n" +
                           str(counter) + " files copied.\r\n")
                file.close()

                self.statusText.insert(tk.END, "\nKLAR!\n")
                self.statusText.insert(
                    tk.END, f'\n{counter} bilder skapade!')
                self.statusText.insert(
                    tk.END, f'\n{failCounter} celler tomma och ignorerades')
            except PermissionError:
                self.statusText.insert(
                    tk.END, "ERROR: Vänligen stäng " + self.excel_file)
            except Exception as error:
                self.statusText.insert(tk.END, error)

    def openExcelFile(self):
        try:
            os.startfile(self.excel_file_n_path)
        except Exception as e:
            self.statusText.insert(e, tk.END)


class Main(tk.Tk):
    def __init__(self):
        super().__init__()

        self.geometry('410x400')
        self.title("Copy Rename 4.0 ADMIN")
        self.resizable(False, False)
        self.attributes('-alpha', 0.92)


if __name__ == '__main__':
    run = Main()
    mainFrame(run)
    run.mainloop()
