import os
import shutil
from tkinter.filedialog import *
import tkinter
import openpyxl
from datetime import datetime
import re
from getkey import getkey, keys

class Main():
    def __init__(self):
        self.current_dir = os.getcwd()
        self.wb = openpyxl.load_workbook('copy-data.xlsx')
        self.sheet = self.wb['Blad1']
        
    def run(self):
        self.mainMenu()
    
    def mainMenu(self):
        self.menu = ['Kontrollera bilder\t', 'Kör\t\t\t', 'Avsluta\t\t\t']
        self.menu_selection = 0
        
        while True:
            self.clear_term()
            print('*'*32)
            print('*        Copy Rename 2.0       *')
            print('*'*32)
            print()
            if self.menu_selection == 0:
                print(self.menu[0] + '<--')
                print(self.menu[1])
                print(self.menu[2])
            if self.menu_selection == 1:
                print(self.menu[0])
                print(self.menu[1] + '<--')
                print(self.menu[2])
            if self.menu_selection == 2:
                print(self.menu[0])
                print(self.menu[1])        
                print(self.menu[2] + '<--')        
            
            keyPressed = getkey()
            if keyPressed == keys.DOWN and self.menu_selection != len(self.menu) -1:
                self.menu_selection += 1
            if keyPressed == keys.UP and self.menu_selection != 0:
                self.menu_selection -= 1
            if keyPressed == keys.ENTER:
                if self.menu_selection == 0:
                    print()
                    print(self.menu[0])
                    self.image_check()
                    input("\nTryck ENTER för att fortsätta...")
                    
                if self.menu_selection == 1:
                    print()
                    print(self.menu[1])
                    self.copy_and_rename()
                    input("\nTryck ENTER för att fortsätta...")
                    
                if self.menu_selection == 2:
                    input("\nProgrammet avslutas...")
                    break                       
        
    def clear_term(self):
        if os.name == 'posix':
            _ = os.system('clear')
        _ = os.system('cls')
        
    def openWriteFile(self, new_file, from_source):
        now = datetime.now()
        time = now.strftime("%Y-%m-%d, %H:%M:%S")
        file = open("logfile.txt", "a+")
        file.write(time + ": " + "Created: " + new_file + " From: " + from_source +"\n")  
        file.close()  
        
    def image_check(self):
        print("Checking source files vs Excel file...")
        
        missing_files = []
        path_list = []
        files_in_dir_list = []
        
        for row in self.sheet.iter_rows(min_row=2, max_col=3, values_only=True):
            files_in_dir = os.listdir(row[0])
            for file in files_in_dir:
                files_in_dir_list.append(file)
            for pic in range(1,3):
                path_list.append(row[pic])
            for i in path_list:
                if i not in files_in_dir_list and not i == None:
                    missing_file_path = row[0]
                    missing_files.append(i)
              
        
        if len(missing_files) <= 0:
            print("All files accounted for! Proceeding...")

        else:
            missing_files_set = set(missing_files)
            missing_files_unique = list(missing_files_set)
            print("Missing", len(missing_files_unique), "source files:", missing_files_unique) 
            
    
    def copy_and_rename(self):
        folder_n_path = askdirectory(title='--------Select Destination Folder')

        try:
            makeFolder = os.makedirs(folder_n_path)
            print("Destination folder missing, creating...")
        except:
            print("Destination folder located, duplicating files...")

        counter=0
        failCounter = 0

        #Logfile.txt
        now = datetime.now()
        time = now.strftime("%Y-%m-%d, %H:%M:%S")
        file = open("logfile.txt", "a+")
        file.write("-----------------------------------------\n")
        file.write("Job started at " + time + "\n")
        file.write("-----------------------------------------\n")
        file.close()

        import time
        startTime = time.time()
        for row in self.sheet.iter_rows(min_row=2, max_col=4, values_only=True):
            for char in row:
                if type(char) is str:
                    fileExtention = char[-4:]

            #Copying the primary picture and renaming it according the the Excel file.
            file_name_front = str(row[3])+"_h1"+str(fileExtention)
            source_file_n_path_front = os.path.join(row[0], str(row[1]))
            file_n_path_destination_front = os.path.join(folder_n_path, file_name_front)
            if row[1] != None:
                shutil.copy(source_file_n_path_front, file_n_path_destination_front)
                print(file_n_path_destination_front)
                counter+=1
                self.openWriteFile(str(file_n_path_destination_front), str(source_file_n_path_front))
            else:
                failCounter+=1
                continue

            #Copying the secondary picture and renaming it according the the Excel file.      
            file_name_back = str(row[3])+"_h2"+str(fileExtention)
            source_file_n_path_back = os.path.join(row[0], str(row[2]))
            file_n_path_destination_back = os.path.join(folder_n_path, file_name_back)
            if row[2] != None:
                shutil.copy(source_file_n_path_back, file_n_path_destination_back)
                print(file_n_path_destination_back)
                counter+=1
                self.openWriteFile(str(file_n_path_destination_back), str(source_file_n_path_back))
            else:
                failCounter+=1
                continue
        time = now.strftime("%Y-%m-%d, %H:%M:%S")
        file = open("logfile.txt", "a+")
        file.write(f"Job complete {time}\n{counter} files copied.\r\n")        
        file.close()
        import time
        print("Done!")  
        print(counter, "files duplicated and renamed!")
        print(failCounter, "empty cell in Excel file, therefore no action")
        endTime = time.time()
        jobTime = (endTime - startTime)
        print(f"Job complete in: {jobTime:.2f} sek")
    

if __name__ == '__main__':
    root = tkinter.Tk()
    root.withdraw()
    run = Main()
    run.run()