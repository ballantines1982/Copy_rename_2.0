import tkinter as tk
from tkinter import ttk
from datetime import datetime
from tkinter.font import BOLD
from openpyxl import load_workbook
import win32com.client


FORMAT = "%H.%M.%S"

def cur_time():
    now = datetime.time(datetime.now())
    now2 = now.strftime(FORMAT)
    current_time = datetime.strptime(now2, FORMAT)
    
    return current_time


def right_col():
    now = datetime.now()
    today = datetime.isoweekday(now)
    if today == 1 or today >5:
        return 1
    else:
        return today+today-1
        return 9

outlook = win32com.client.Dispatch('outlook.application')
mapi = outlook.GetNamespace("MAPI")
sent = mapi.GetDefaultFolder(5)
messages = sent.Items

# for mail in messages:
#     print(mail)


#TODO: Lägg till en rad som räknar ut vilken vecka samt veckodag leveransen kommer

cell_lst= []

wb = load_workbook('Beställningsschema.xlsx')
sheet = wb['Blad1']

start_col = right_col()

name_lst = []
time_lst = []
prefix_lst = []

for row in sheet.iter_rows(min_row=3, min_col=start_col, max_col=start_col+1 ,max_row=24 ,values_only=True):
    for cell in row:
        cell_lst.append(cell)
print(cell_lst)        
for i in range(len(cell_lst)):
    if i % 2 == 0 and not i % 4 == 0:
        name_lst.append(cell_lst[i])
    elif i % 4 == 0:
        prefix_lst.append(cell_lst[i])
    elif cell_lst[i] != None and "." in cell_lst[i]:
        t = cell_lst[i].split(' ')
        time_lst.append(t[0])
print(len(time_lst), time_lst)
print(len(name_lst), name_lst)
print(len(prefix_lst), prefix_lst)



class Supplier(tk.Frame):
    def __init__(self, container, name, time, prefix="", mail=""):
        super().__init__(container)
        self.name = name
        self.time = datetime.strptime(time + ".00", FORMAT)
        self.prefix = prefix
        self.mail = mail
        self.cur_time = cur_time()
        self.diff = self.time - self.cur_time
        
        self.columnconfigure(0, weight=3)
        self.columnconfigure(1, weight=1)

        self.sup_name = tk.Label(self, text=str(self.name), bg='lightgrey')
        self.sup_name.grid(column=0, 
                                row=0, 
                                padx=2, 
                                pady=3, 
                                sticky=tk.W)
        
        self.sup_time = tk.Label(self, text=str(self.diff), bg='lightgrey', font=('Arial', 14, BOLD))
        self.sup_time.grid(column=1,
                                row=0,
                                padx=2,
                                pady=3,
                                sticky=tk.E)
        
        self._update()

    def _update(self):
        
        self.cur_time = cur_time()
        self.diff = self.time - self.cur_time
        if self.diff.days < 0:
            self.sup_time.configure(text= "OVERDUE")
        else:
            self.sup_time.configure(text = self.diff)
        
            self.sup_time.after(1000, self._update)
        



class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.geometry('250x600')
        self.resizable(False, False)
        self.frm_lst = [] 
        
        self.columnconfigure(0, weight=1)
        
        self._create_widgets()
 
        
    def _create_widgets(self):
        
        for i in range(len(time_lst)):
            self.frame = Supplier(self, name_lst[i], time_lst[i])
            self.frame['borderwidth'] = 1
            self.frame['relief'] = 'raised'
            self.frame.config(bg='lightgrey')
            #self.frame['padding'] = (10,10,10,0)
            self.frame.grid(column=0, row=i, padx=10, pady=5, sticky=tk.EW)
            self.frm_lst.append(self.frame)
        
        self.frame.sup_time.after(1000, self.frame._update)
                


if __name__ == '__main__':
    run = App()
    run.mainloop()