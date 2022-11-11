from openpyxl import load_workbook
import pandas as pd
import datetime
#from dateutil import parser
import re
import locale

locale.setlocale(locale.LC_ALL, 'Swedish_Sweden.1252')

date_str = "25:e Maj 2022"

def char_search(string): 
    if re.search(':e', string):
        string = re.sub(':e', '', string)
        return string
    return string

try:
    date_obj = datetime.datetime.strptime(char_search(date_str), '%d %b %Y').date()
    print(date_obj)
except Exception as e:
    print(e)             
                        
                        
                        
def new_sheetsNames(start=0, end=0):
    file = r'C:\Users\svard\OneDrive\Skrivbord\Privat_budget_210921.xlsx'
    newfile = r'C:\Users\svard\OneDrive\Skrivbord\Privat_budget_220703.xlsx'

    wb = load_workbook(file)

    startYear = 2020
    months = ['Januari', 'Februari', 'Mars', 'April', 'Maj', 'Juni', 'Juli', 'Augusti', 'September', 'Oktober', 'November', 'December']
    
    for sheet in wb.sheetnames:
        _sheet = wb[sheet]
        if '25e' in sheet:
            
            for j in range(len(months)): 
                if str(months[j]) in sheet:
                    
                    date = f'25e {months[j]} {str(startYear)}'
                    _sheet.title = date
                    wb.save(newfile)
                    if j == len(months)-1:
                        startYear += 1
                        j = 0
            
    print('Alla kolumnnamn ändrade...')