import sqlite3
from openpyxl import load_workbook
from sqlite3.dbapi2 import Error
from tkinter.filedialog import *
import os

#Bravida netto, 0,1,4
#mascot 0,1,5,5

class NewDBTable:
    def __init__(self, table):
        self.con = sqlite3.connect('art_db.db')
        self.cur = self.con.cursor()
        self.table = table
        
        
    def loadDb(self):
        if self.con:
            print('Connected to database...')
            self.cur.execute("""CREATE TABLE IF NOT EXISTS {} (
                                artnr integer NOT NULL UNIQUE, 
                                art_name text NOT NULL, 
                                cost NOT NULL, 
                                sale float NOT NULL)""".format(self.table))
        else:
            print('Error when trying to connect to database..')
            
            
    def excelFile(self, file, sheet, mincol, c1, c2, c3, c4):
        self.wb = load_workbook(file)
        self.sheet = self.wb[sheet]
        MAXROW = self.sheet.max_row
        counter = 0
        
        sql = '''INSERT INTO {} (artnr, art_name, cost, sale) VALUES(?,?,?,?)'''.format(self.table)
        temp_list = []
        for row in self.sheet.iter_rows(min_row=mincol, max_row=MAXROW, values_only=True):
            artnr = row[c1]
            description = row[c2]
            cost = float(row[c3])
            sale = float(row[c4])

            try:
                temp_list.append(artnr)
                temp_list.append(description)
                temp_list.append(cost)
                temp_list.append(sale)
                self.cur.execute(sql, temp_list)
            except sqlite3.IntegrityError:
                pass
            else:
                self.con.commit()
                counter+=1
            temp_list = []
        print(f"{counter} rows injected to DB..")
        

if __name__ == '__main__':
    mascot = NewDBTable('Mascot')
    mascot.loadDb()
    mascot.excelFile('Pricelist_Januar2022_SE.xlsx','1400_01_2022_SV_20210927',2,0,2,5,5)
    
    
