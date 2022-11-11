import win32com.client
import openpyxl as pyxl
from openpyxl import load_workbook
import inspect

# try:
#     wb = load_workbook("PO.xlsx")
#     ws = wb.active
#     print("Source file found...")
# except Exception as e:
#     print(e)
    
    

outlook = win32com.client.Dispatch("Outlook.application")
mapi = outlook.GetNameSpace("MAPI")

# for account in mapi.Accounts:
#     print(account.DeliveryStore.DisplayName)
    
inbox = mapi.GetDefaultFolder(6)
message = inbox.Items


for i in inspect.getmembers(message):
    print(i)











# if len(message) == 0:
#     print("No emails found...")
# else:
#     print(f"Reading {len(message)} email...")

# for item in message:            
#     for row in range(2, ws.max_row+1):
#         for col in range(1, ws.max_column+1):
#             checkCell = ws.cell(row=row, column=3).value
#             cellValue = ws.cell(row=row, column=col).value
#             searchWord = ws.cell(row=row, column=2).value
            
#             try:
#                 if cellValue != None:
#                     if searchWord in item.Subject:
#                         ws.cell(row=row, column=4).value = str(item.ReceivedTime.date())
#                         print(f"{item.ReceivedTime.date()}: {cellValue} in {item}")
#                         checkCell = "X"
#             except Exception as e:
#                 print(e)
    

# try:
#     wb.save('PO.xlsx')
#     print("File saved...")
# except Exception as e:
#     print(e)
    