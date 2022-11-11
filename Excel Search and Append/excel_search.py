import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os


source_path = os.getcwd()
source_file = 'source.xlsx'
source_file_n_path = os.path.join(source_path, source_file)

target_file = 'target.xlsx'
target_file_n_path = os.path.join(source_path, target_file)

target = load_workbook(target_file_n_path)
Tsheets = target.sheetnames
target_ws = target[Tsheets[0]]

source = load_workbook(source_file_n_path)
Ssheets = source.sheetnames
source_ws = source[Ssheets[0]]
lst = []
for i in range(1, source_ws.max_row +1):
    for j in range(1, source_ws.max_column +1):
        if source_ws.cell(row=i, column=j).value == 660006:
            lst.append(source_ws.cell(row=i, column=j).value)

print(lst)