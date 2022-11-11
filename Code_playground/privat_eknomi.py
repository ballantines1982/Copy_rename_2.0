import openpyxl
from openpyxl import load_workbook
import os
from tkinter.filedialog import askopenfilename
import pandas as pd
import numpy


file = r'C:\Users\svard\OneDrive\Skrivbord\Privat_budget_210921.xlsx'

wb = load_workbook(filename=file, data_only=True, read_only=True)
sheet = wb.worksheets[-3]

def average(cell):
    lst = []
    months = []
   
    for sheet in wb.worksheets:
        if "25e " in sheet.title:
            print(sheet.title + " = " + str(sheet[cell].value) + "kr")
            lst.append(int(sheet[cell].value))
            months.append(sheet.title)
    lst_sum = sum(lst)
    average = lst_sum / len(lst)
    list_max = max(lst)
    list_min = min(lst)
    max_index = lst.index(list_max)
    min_index = lst.index(list_min)
    print("-----------------------------")
    print(f"Cell average: {average:.0f}kr")
    print(f"Highest month: {months[max_index]}, {list_max}kr")
    print(f"Lowest month: {months[min_index]}, {list_min}kr")

def saving(cells):
    #cells = ["Spara Bilen", "Spara Familjen", "Spara Helgnöje"]
    lst = []
    for sheets in wb.worksheets:
        if "25e" in sheets.title:
            for row in sheets.iter_rows(max_col=4, values_only=True):
                for search in cells: 
                    if search in row:
                        lst.append(row)
    slim_lst = []
    for i in range(4):
        for search in lst:
            print(search)
            if search in lst:
                slim_lst.append(search[i])
                slim_lst.append(search[i+1])
    print(slim_lst)



#average('B20')
saving(["Spara Bilen", "Spara Familjen", "Spara Helgnöje"])