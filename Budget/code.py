from json import load
import sqlite3 as sql
from turtle import color
from openpyxl import load_workbook
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np



wb = load_workbook(r'C:\Users\svard\OneDrive\Skrivbord\Privat_budget_210921.xlsx', data_only=True)

p_income = []
s_income = []
tot_income = []

for sheet in wb.worksheets:
    if '25' in sheet.title:
        # print('-'*100)
        # print(sheet)
        # print('-'*50)
        for row in sheet.iter_rows(max_col=4, max_row=2, min_row=2, values_only=True):
            #print(row)
            tot_sum = row[1] + row[3]
            p_income.append(row[1])
            s_income.append(row[3])
            tot_income.append(tot_sum)
            
            
print(len(p_income))
print(p_income)
months = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]

plt.plot(months, p_income, color='green', label="Peter")
plt.plot(months, s_income, color='lightgreen', label="Sara")
plt.plot(months, tot_income, color="blue", label='Summerad inkomst')
plt.ylabel('Inkomst')
plt.xlabel('Månader')
plt.title("Hjertsdotter Svärd Ekonomi")
plt.axis([0,15,5000,90000])
plt.grid(True)
plt.legend()
plt.show() 


# con = sql.connect('test-db.db')
# cur = con.cursor()
# peter_table = """ CREATE TABLE IF NOT EXISTS budget_peter (
#     id integer PRIMARY KEY,
#     bill_name text, 
#     bill_value integer
#     ); """
    
# sara_table = """ CREATE TABLE IF NOT EXISTS budget_sara (
#     id integer PRIMARY KEY,
#     bill_name text, 
#     bill_value integer
#     ); """

# for table in [peter_table, sara_table]:    
#     cur.execute(table)
    
# cur.execute("SELECT name FROM sqlite_master WHERE type='table';")
# print(cur.fetchall())

# con.close()