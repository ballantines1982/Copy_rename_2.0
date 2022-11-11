import openpyxl
import os
from openpyxl.utils import *
import time

from openpyxl.xml.constants import MAX_COLUMN

os.chdir(r'C:\Python Projekt\Copy Rename Project\training')
wb_source = openpyxl.load_workbook('data.xlsx')

def find_sheet(search):
    for i in range(len(wb_source.sheetnames)):
        if search in wb_source.sheetnames[i]:
            return i

sheet = wb_source[wb_source.sheetnames[find_sheet('Artiklar')]]

search_list = ['Kund', 'SE10', 'MOQ']

data_list = []
column_list = []
column = None
col_num = 0

# ######################################################################################################################


startTime = time.time()

for col in sheet.iter_cols(min_row=4, max_row=4, values_only=True):
    col_num += 1
    for search_word in search_list:
        if search_word in col:
            column_list.append(col_num)
            
print("column_list: " + str(column_list))


endTime = time.time()
jobTime = (endTime - startTime)
print(f"First loop complete in: {jobTime:.2f} sek")

# #######################################################################################################################


for row in sheet.iter_rows(min_row=5, max_row=20, values_only=True):
    #print(row)
    data_list.append(row[4-1])
    data_list.append(row[13-1])
    data_list.append(row[73-1])

            

print(data_list)
print(len(data_list))
print(column_list)        
        
        
