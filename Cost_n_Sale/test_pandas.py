
import pandas as pd
import sqlite3
from openpyxl import load_workbook



class Main():
    def __init__(self):
        try:
            self.con = sqlite3.connect('test_db.db')
        except sqlite3.Error as e:
            print(e)
        else:
            self.wb = pd.read_excel('Fristads2022.xlsx', sheet_name=None)
            for sheet in self.wb:
                try:
                    self.wb[sheet].to_sql(sheet, self.con, index=False)
                except ValueError as e:
                    print(e)
        self.con.commit()
        self.renameColumns()
        self.con.close()        
                
                
    def renameColumns(self):
        wb = load_workbook('Fristads2022.xlsx')
        sheet = wb.active
        col_headers = []
        for row in sheet.iter_rows(min_row=13, max_row=13, min_col=2, values_only=True):
            col_headers.append(row)
        
        self.con = sqlite3.connect('test_db.db')
        
        for i in range():
        
        
        
        sql = """ALTER TABLE [ 2022 Scandinavian range] RENAME COLUMN [Unnamed: 5] TO Supplier"""            
        try:
            self.cur = self.con.cursor()
            self.cur.execute(sql)
            print('Connected...')
        except sqlite3.Error as e:
            print(e)
        else:
            print('Column renamed...')
            
            
            
if __name__ == '__main__':
    run = Main()