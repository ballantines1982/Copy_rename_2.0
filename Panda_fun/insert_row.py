from re import L
import openpyxl
from openpyxl import load_workbook

file = 'Privat_budget_210921 - 3 - modified.xlsx'
newfile = 'Privat_budget_210921 - 3 - row added.xlsx'

wb = load_workbook(file)
flag = False
for sheet in wb.sheetnames:
    wh = wb[sheet]
    if '20' in sheet:
        wh.insert_rows(20, amount=3)
        wb.save(newfile)
