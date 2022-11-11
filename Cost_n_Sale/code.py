from sqlite3.dbapi2 import Error
from tkinter.font import NORMAL
from openpyxl import load_workbook
import sqlite3
import tkinter as tk
from tkinter import ttk
import time


class Main(tk.Tk):
    def __init__(self):
        super().__init__()
        self.geometry('800x500+400+200')
        self.title("Cost & Sale")

        self.create_widgets()

    def create_widgets(self):
        mainframe = MainFrame(self)
        mainframe.grid(row=0, column=0)


class MainFrame(ttk.Frame):
    def __init__(self, container):
        super().__init__(container)
        # self.excelFile = ExcelFile()
        self.dbObject = ExcelFile()
        
        self.create_widgets()

    def create_widgets(self):
        ttk.Label(self, text='Databas').grid(row=0, column=0)
        
        self.var1 = tk.IntVar()
        self.searchEntry = ttk.Entry(self, textvariable=self.var1)
        self.searchEntry.grid(row=1, column=0)
        self.searchEntry.focus()
        
        self.selected_cust = tk.StringVar()
        self.cust_select = ttk.Combobox(self, textvariable=self.selected_cust)
        self.cust_select['values'] = ['Bravida', 'NOS', 'Ingen kund']
        self.cust_select['state'] = 'readonly'
        self.cust_select.set('Ingen kund')
        self.cust_select.grid(row=1, column=1)
        
        self.searchBtn = ttk.Button(self, text='Search', command=self.searchDB)
        self.searchBtn.grid(row=1, column=2)
        
        self.updateBtn = ttk.Button(self, text='Reload DB', command=self.dbObject.read_excel)
        self.updateBtn.grid(row=1, column=3)
        
        ttk.Separator(self, orient='horizontal').grid(row=2, column=0, columnspan=4, sticky='EW', pady=5)
        
        self.DBcolumns = ('Artnr', 'Benämning', 'Inpris', 'Utpris')
        self.DBtree = ttk.Treeview(self, columns=self.DBcolumns, show='headings')
        for i in self.DBcolumns:
            self.DBtree.heading(i, text=i)
        self.DBtree.column('Artnr', width=100, anchor=tk.W)
        self.DBtree.column('Benämning', width=225, anchor=tk.W)
        self.DBtree.column('Inpris', width=70, anchor=tk.W)
        self.DBtree.column('Utpris', width=70, anchor=tk.W)
        self.DBtree.grid(row=3, column=0, columnspan=4)
        
        self.box = ttk.Checkbutton(self, text='Bravida', command=self.bravidaCol, offvalue=0, onvalue=1)
        self.box['state'] = NORMAL
        self.box.grid(row=3, column=5)
        
    def bravidaCol(self):
        pass    
        
    def searchDB(self):
        startTime = time.time()
        for i in self.DBtree.get_children():
            self.DBtree.delete(i)
        artnr = ('%' + self.searchEntry.get() + '%')
        
        q = """SELECT * FROM fristads 
               WHERE fristads.art_name LIKE ? OR fristads.artnr LIKE ? 
               UNION 
               SELECT * FROM Mascot 
               WHERE Mascot.art_name LIKE ? OR Mascot.artnr LIKE ? 
               ORDER BY cost ASC"""
               
        self.con = DataBase()
        self.cur = self.con.db.cursor()
        try:
            self.cur.execute(q, [artnr, artnr, artnr, artnr])
        except Error as e:
            print(e)
        else:
            rows = self.cur.fetchall()
            for row in rows:
                data = row[0], row[1], round(row[2], 2), round(row[3], 2)
                self.DBtree.insert('', tk.END, values=data)
                if len(rows) == 0:
                    self.DBtree.insert('', tk.END, text="Inga träffar...")
            endTime = time.time()
            jobTime = (endTime - startTime)
            print(f"Search complete in: {jobTime:.2f} sek")
             
            
class DataBase():
    def __init__(self):
        self.db = sqlite3.connect('art_db.db')
        if self.db:
            print('Connected to database...')
            self.cur = self.db.cursor()
            self.cur.execute('''CREATE TABLE IF NOT EXISTS fristads(
                artnr integer NOT NULL UNIQUE,
                art_name text NOT NULL, 
                cost float NOT NULL,
                sale float NOT NULL);''')
        else:
            print('Error when trying to connect to database..')



class ExcelFile(DataBase):
    def __init__(self):
        self.excel_data = []
        self.DB = DataBase()
        
    def read_excel(self):
        try:
            startTime = time.time()
            self.wb = load_workbook('FRISTADS PRISFIL INK. STATNR. SEK 1 FEB 2022 - SE.xlsx')
        except Error as fee:
            print(fee)
        else:
            endTime = time.time()
            jobTime = (endTime - startTime)
            print(f'Excel file found in {jobTime:.2f} sek')
            self.sheet = self.wb.active
        print('Reading rows..')
        sql = '''INSERT INTO fristads(artnr, art_name, cost, sale) VALUES(?,?,?,?)'''
        temp_list = []
        #self.DB = DataBase()
        # self.db = sqlite3.connect('art_db.db')
        self.cur = self.DB.db.cursor()
        counter = 0
        
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            artnr = row[0]
            description = row[6]
            cost = float(row[19]) * 0.427
            sale = float(row[19])
            # sale = float(row[19]) * 0.8 
            try:
                temp_list.append(artnr)
                temp_list.append(description)
                temp_list.append(cost)
                temp_list.append(sale)
                self.cur.execute(sql, temp_list)
            except sqlite3.IntegrityError:
                pass
                #print("Data finns redan, skippar och fortsätter..")
            else:
                self.DB.db.commit()
                counter += 1
            temp_list = []
        print(f"{counter} rows injected to DB..")
    


if __name__ == "__main__":
    app = Main()
    app.mainloop()

    
    
